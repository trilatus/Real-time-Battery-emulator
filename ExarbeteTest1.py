import math
import matplotlib.pyplot as plt
import openpyxl

PARAMETERS = {"soc":[1,0.979455857856746,0.959189339255969,0.939200444197668,0.919211549139367,0.899222654081066,0.849250416435314,0.799278178789561,0.749305941143809,0.699611327040533,0.649639089394781,0.599666851749028,0.549694614103276,0.499722376457524,0.449750138811771,0.399777901166019,0.349528039977790,0.299555802332038,0.249583564686285,0.199611327040533,0.149639089394781,0.0999444752915047,0.0796779566907274,0.0599666851749028,0.0399777901166019,0.0199888950583009,0],"ocv":[4.20059028171202,4.16445653203492,4.14012975447597,4.11793510998830,4.09733423488953,4.07980521404167,4.03471248342643,3.98866453343482,3.93119212338236,3.87352922950037,3.83465671865652,3.79917203042368,3.78245430078528,3.75949626563702,3.73467566916230,3.71464781193304,3.68808193297151,3.66808619543555,3.63583981697506,3.59994730963900,3.57430273447701,3.54612302249662,3.49076110025690,3.42024513850162,3.31767959768643,3.17747531909001,3.01476509201849],"r0":[0.00719146897428723,0.00741850900708751,0.00743403420495271,0.00748776198313241,0.00721529818378370,0.00716371327631248,0.00668149676770256,0.00685037715211540,0.00687085756080030,0.00704421277070143,0.00658139870114716,0.00602943743802638,0.00604826154895946,0.00633867560153846,0.00649461462617092,0.00654641416094327,0.00692275952370252,0.00727251528016254,0.00742656026487787,0.00789437433539946,0.0111292432537967,0.0153515136996802,0.0200853634227497,0.0237805732556751,0.0300469269478585,0.0399182691648787,0.0656593972614554],"r1":[0.00518399051058848,0.00445762771780047,0.00445762771780047,0.00445762771780047,0.00445762771780047,0.00445762771780047,0.00445762771780047,0.00445762771780047,0.00445762771780047,0.00445762771780047,0.00445762771780047,0.00445762771780047,0.00445762771780047,0.00445762771780047,0.00445762771780047,0.00445762771780047,0.00445762771780047,0.00445762771780047,0.00445762771780047,0.00445762771780047,0.00445762771780047,0.00445762771780047,0.00445762771780047,0.00445762771780047,0.00445762771780047,0.00445762771780047,0.00445762771780047],"r2":[0.00437254212211296,0.00402688525769237,0.00402688525769237,0.00402688525769237,0.00402688525769237,0.00402688525769237,0.00402688525769237,0.00402688525769237,0.00402688525769237,0.00402688525769237,0.00402688525769237,0.00402688525769237,0.00402688525769237,0.00402688525769237,0.00402688525769237,0.00402688525769237,0.00402688525769237,0.00402688525769237,0.00402688525769237,0.00402688525769237,0.00402688525769237,0.00402688525769237,0.00402688525769237,0.00402688525769237,0.00402688525769237,0.00402688525769237,0.00402688525769237],"r3":[0.00588621223605553,0.0158863088300064,0.0158863088300064,0.0158863088300064,0.0158863088300064,0.0158863088300064,0.0158863088300064,0.0158863088300064,0.0158863088300064,0.0158863088300064,0.0158863088300064,0.0158863088300064,0.0158863088300064,0.0158863088300064,0.0158863088300064,0.0158863088300064,0.0158863088300064,0.0158863088300064,0.0158863088300064,0.0158863088300064,0.0158863088300064,0.0158863088300064,0.0158863088300064,0.0158863088300064,0.0158863088300064,0.0158863088300064,0.0158863088300064],"t1":[9.86264106192937,9.05625483038048,9.05625483038048,9.05625483038048,9.05625483038048,9.05625483038048,9.05625483038048,9.05625483038048,9.05625483038048,9.05625483038048,9.05625483038048,9.05625483038048,9.05625483038048,9.05625483038048,9.05625483038048,9.05625483038048,9.05625483038048,9.05625483038048,9.05625483038048,9.05625483038048,9.05625483038048,9.05625483038048,9.05625483038048,9.05625483038048,9.05625483038048,9.05625483038048,9.05625483038048],"t2":[200.107655784865,195.293652708401,195.293652708401,195.293652708401,195.293652708401,195.293652708401,195.293652708401,195.293652708401,195.293652708401,195.293652708401,195.293652708401,195.293652708401,195.293652708401,195.293652708401,195.293652708401,195.293652708401,195.293652708401,195.293652708401,195.293652708401,195.293652708401,195.293652708401,195.293652708401,195.293652708401,195.293652708401,195.293652708401,195.293652708401,195.293652708401],"t3":[637.560091248790,701.928692420017,701.928692420017,701.928692420017,701.928692420017,701.928692420017,701.928692420017,701.928692420017,701.928692420017,701.928692420017,701.928692420017,701.928692420017,701.928692420017,701.928692420017,701.928692420017,701.928692420017,701.928692420017,701.928692420017,701.928692420017,701.928692420017,701.928692420017,701.928692420017,701.928692420017,701.928692420017,701.928692420017,701.928692420017,701.928692420017]}
BAT_AH = 10.00555556
BAT_LOAD = 10

class Params:
    def __init__(self, params, bat_ah, bat_load):
        self.bat_ah, self.bat_load = bat_ah, bat_load
        self.soc, self.ocv = params["soc"], params["ocv"]
        self.r0, self.r1, self.r2, self.r3 = params["r0"], params["r1"], params["r2"], params["r3"]
        self.t1, self.t2, self.t3 = params["t1"], params["t2"], params["t3"]
        self.step_calc()

    def step_calc(self):
        res = []
        time_steps = []
        dis_speed = self.bat_ah/(self.bat_load/3600)
        soc_loss = (self.soc[0] - self.soc[-1])/dis_speed

        for soc_level in self.soc:
            temp_soc = self.soc[0]
            counter = 0
            abs1 = abs(soc_level - temp_soc)

            while temp_soc > 0:
                abs2 = abs(soc_level - temp_soc)
                if abs2 > abs1:
                    res.append(counter)
                    time_steps.append(counter - 1)
                    break
                abs1 = abs2
                counter += 1
                temp_soc -= soc_loss
        
        for i in range(len(res) - 1):
            res[i] = res[i + 1] - res[i]
        res[-1] = round(dis_speed) - res[-1] + 1

        time_steps.pop(0)
        time_steps.append(round(dis_speed))

        self.soc_loss = soc_loss
        self.steps = res
        self.discharge_speed = dis_speed
        self.time_steps = time_steps

class Sim:
    def __init__(self, params, excel = False, pulses = 0, rest_time = 0):
        self.excel = excel
        if self.excel: self.excel_compare()
        self.params = params
        self.discharge_time = 0
        self.load_time = 0
        self.load = False
        self.pulses = pulses
        self.rest_time = rest_time
        self.timeline_setup()
        self.voltage = []
        self.time = []

    def vl_calc(self):
        idx = 0
        for i in self.params.time_steps:
            if self.discharge_time >= i:
                idx += 1
                if idx == len(self.params.time_steps): idx -= 1
            else: break
        
        instep_time = self.params.steps[idx] - (self.params.time_steps[idx] - self.discharge_time)

        ocv = self.params.ocv[idx] - (self.params.ocv[idx] - self.params.ocv[idx + 1]) / self.params.steps[idx] * instep_time
        r1 = self.params.r1[idx] - (self.params.r1[idx] - self.params.r1[idx + 1]) / self.params.steps[idx] * instep_time
        r2 = self.params.r2[idx] - (self.params.r2[idx] - self.params.r2[idx + 1]) / self.params.steps[idx] * instep_time
        r3 = self.params.r3[idx] - (self.params.r3[idx] - self.params.r3[idx + 1]) / self.params.steps[idx] * instep_time
        t1 = self.params.t1[idx] - (self.params.t1[idx] - self.params.t1[idx + 1]) / self.params.steps[idx] * instep_time
        t2 = self.params.t2[idx] - (self.params.t2[idx] - self.params.t2[idx + 1]) / self.params.steps[idx] * instep_time
        t3 = self.params.t3[idx] - (self.params.t3[idx] - self.params.t3[idx + 1]) / self.params.steps[idx] * instep_time
        load_time = self.load_time

        if self.load:
            r0 = self.params.r0[idx] - (self.params.r0[idx] - self.params.r0[idx + 1]) / self.params.steps[idx] * instep_time
            load = self.params.bat_load
            u0 = load * r0 
            u1 = load * r1 * (1 - math.exp(-load_time/t1))
            u2 = load * r2 * (1 - math.exp(-load_time/t2))
            u3 = load * r3 * (1 - math.exp(-load_time/t3))
            vl = ocv - u0 - u1 - u2 - u3
            return vl
        else:
            u1_nl = r1 * (math.exp(-load_time/t1))
            u2_nl = r2 * (math.exp(-load_time/t2))
            u3_nl = r3 * (math.exp(-load_time/t3))
            vl_nl = ocv - u1_nl - u2_nl - u3_nl
            return vl_nl
 
    def timeline_setup(self):
        self.timeline = []
        if not self.excel:
            dis_speed = round(self.params.bat_ah/(self.params.bat_load/3600))
            if self.pulses == 0:
                for i in range(int(dis_speed)):
                    self.timeline.append(1)
            else:
                for i in range(self.pulses):
                    for i in range(self.rest_time):
                        self.timeline.append(0)
                    for i in range(int(dis_speed/self.pulses)):
                        self.timeline.append(1)
        else:
            self.timeline = self.xl_current
            for i in range(len(self.timeline)):
                if self.timeline[i] != 0: self.timeline[i] = 1
            
    def run(self):
        for i in range(len(self.timeline)):
            if self.timeline[i] == 1:
                if not self.load:
                    self.load_time = 0
                    self.load = True
                else: self.load_time += 1
                self.discharge_time += 1
            else:
                if self.load:
                    self.load_time = 0
                    self.load = False
                else: self.load_time += 1

            self.voltage.append(self.vl_calc())
            self.time.append(i)

    def plot(self):
        plt.plot(self.time, self.voltage, label = "Sim")
        if self.excel: plt.plot(self.time, self.xl_voltage, label = "Real")
        plt.legend()
        plt.show()
    
    def excel_compare(self):
        wb_name = "test_data.xlsx"
        wb = openpyxl.load_workbook(wb_name)
        sheet = wb.active
        self.xl_time = []
        self.xl_current = []
        self.xl_voltage = []
        for i in range(2, sheet.max_row + 1):
            self.xl_time.append(int(sheet.cell(row = i, column = 1).value))
            self.xl_current.append(int(sheet.cell(row = i, column = 2).value))
            self.xl_voltage.append(float(sheet.cell(row = i, column = 3).value))

parameters = Params(PARAMETERS, BAT_AH, BAT_LOAD)
sim = Sim(parameters, True)

sim.run()
sim.plot()

